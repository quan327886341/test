import openpyxl, os


wb = openpyxl.load_workbook(os.path.join(os.getcwd(), 'produceSales.xlsx'))
sheet = wb.get_sheet_by_name('Sheet')

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.09,
                 'Lemon': 1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')

import PyPDF2
